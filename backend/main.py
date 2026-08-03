from contextlib import asynccontextmanager
from fastapi import FastAPI, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from openpyxl import load_workbook
import io
import os
from typing import Dict, List, Any, Optional
from collections import defaultdict

DEFAULT_DATASET_PATH = os.path.join(
    os.path.dirname(__file__), "data", "default_dataset.xlsx"
)

default_dataset_cache: Optional[Dict[str, Any]] = None


def load_default_dataset() -> None:
    global default_dataset_cache

    if not os.path.exists(DEFAULT_DATASET_PATH):
        default_dataset_cache = None
        return

    try:
        with open(DEFAULT_DATASET_PATH, "rb") as f:
            contents = f.read()

        rows, molecules, products = parse_workbook(contents)
        default_dataset_cache = build_upload_response(rows, molecules, products)
    except Exception as e:
        default_dataset_cache = {
            "success": False,
            "error": f"Failed to load default dataset: {e}",
            "analytics": [],
        }


@asynccontextmanager
async def lifespan(app: FastAPI):
    load_default_dataset()
    yield


app = FastAPI(
    title="MolecuLab Analytics Backend",
    version="1.0.0",
    lifespan=lifespan,
)

# CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)


def compute_cagr(year0: float, year1_fallback: float, year2: float) -> float:
    if year0 > 0 and year2 > 0:
        return ((year2 / year0) ** 0.5 - 1) * 100
    if year1_fallback > 0 and year2 > 0:
        return (year2 / year1_fallback - 1) * 100
    return 0.0


def to_number(value):
    if value is None:
        return 0.0

    try:
        return float(value)
    except (ValueError, TypeError):
        return 0.0


SECTOR_KEYS = ["HOSPITAL", "RETAIL"]


def normalize_sector(value) -> str:
    if not value:
        return "UNKNOWN"
    v = str(value).strip().upper()
    return v if v in SECTOR_KEYS else "UNKNOWN"


INNOVATION_KEY_MAP = {
    "INNOVATIVE BRANDED PRODUCTS": "INNOVATIVE_BRANDED_PRODUCTS",
    "NON ORIGINAL BRANDED PRODUCTS": "NON_ORIGINAL_BRANDED_PRODUCTS",
    "UNBRANDED PRODUCTS": "UNBRANDED_PRODUCTS",
}
INNOVATION_KEYS = list(INNOVATION_KEY_MAP.values())


def normalize_innovation(value) -> str:
    if not value:
        return "UNKNOWN"
    v = str(value).strip().upper()
    return INNOVATION_KEY_MAP.get(v, "UNKNOWN")


def compute_analytics(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    if not rows:
        return []

    molecules = defaultdict(list)

    for row in rows:
        mol = row.get("Molecule List")

        if mol:
            molecules[str(mol)].append(row)

    import math

    # Pre-computation pass — derive normalization bounds from the actual dataset
    _pre_competition: List[int] = []
    _pre_rev_cagr: List[float] = []
    for _mol_rows in molecules.values():
        _brands: set = set()
        for _row in _mol_rows:
            _b = _row.get("International Product")
            if _b:
                _b = str(_b).strip()
                if _b:
                    _brands.add(_b)
        _pre_competition.append(len(_brands))

        _r23 = sum(to_number(r.get("MAT Q2 2023_LCD MNF")) for r in _mol_rows)
        _r24 = sum(to_number(r.get("MAT Q2 2024_LCD MNF")) for r in _mol_rows)
        _r25 = sum(to_number(r.get("MAT Q2 2025_LCD MNF")) for r in _mol_rows)
        if _r23 > 0 and _r25 > 0:
            _pre_rev_cagr.append(((_r25 / _r23) ** 0.5 - 1) * 100)
        elif _r24 > 0 and _r25 > 0:
            _pre_rev_cagr.append((_r25 / _r24 - 1) * 100)
        else:
            _pre_rev_cagr.append(0.0)

    max_competition = max(_pre_competition) if _pre_competition else 1
    max_competition = max(max_competition, 2)  # guard against single-molecule datasets

    cagr_floor = min(_pre_rev_cagr) if _pre_rev_cagr else -50.0
    cagr_cap = 150.0

    def norm_revenue(v: float) -> float:
        if v <= 0:
            return 0.0
        log_rev = math.log10(max(v, 1000))
        return max(0.0, min(1.0, (log_rev - 3.0) / 6.0))

    def norm_cagr(v: float) -> float:
        clamped = max(cagr_floor, min(cagr_cap, v))
        cagr_range = cagr_cap - cagr_floor
        return (clamped - cagr_floor) / cagr_range if cagr_range else 0.5

    def norm_competition(c: int) -> float:
        clamped = max(1, min(max_competition, c))
        return 1 - (clamped - 1) / (max_competition - 1)

    analytics = []

    for molecule, mol_rows in molecules.items():
        revenues_2023 = [
            to_number(r.get("MAT Q2 2023_LCD MNF"))
            for r in mol_rows
        ]

        revenues_2024 = [
            to_number(r.get("MAT Q2 2024_LCD MNF"))
            for r in mol_rows
        ]

        revenues_2025 = [
            to_number(r.get("MAT Q2 2025_LCD MNF"))
            for r in mol_rows
        ]

        stds_2023 = [
            to_number(r.get("MAT Q2 2023_Standard Units"))
            for r in mol_rows
        ]

        stds_2024 = [
            to_number(r.get("MAT Q2 2024_Standard Units"))
            for r in mol_rows
        ]

        stds_2025 = [
            to_number(r.get("MAT Q2 2025_Standard Units"))
            for r in mol_rows
        ]

        rev_2023 = sum(revenues_2023)
        rev_2024 = sum(revenues_2024)
        rev_2025 = sum(revenues_2025)

        std_2023 = sum(stds_2023)
        std_2024 = sum(stds_2024)
        std_2025 = sum(stds_2025)

        std_cagr = compute_cagr(std_2023, std_2024, std_2025)
        rev_cagr = compute_cagr(rev_2023, rev_2024, rev_2025)

        # Single pass over this molecule's rows: brand-level detail (per-year
        # revenue + manufacturer/corporation), sector split, and innovation mix.
        # All three are revenue-share breakdowns of the same 3-year total, kept
        # consistent with Dominance_Ratio's convention.
        brand_detail: Dict[str, Dict[str, Any]] = {}
        sector_revenue: Dict[str, float] = defaultdict(float)
        innovation_revenue: Dict[str, float] = defaultdict(float)

        for row in mol_rows:
            r23 = to_number(row.get("MAT Q2 2023_LCD MNF"))
            r24 = to_number(row.get("MAT Q2 2024_LCD MNF"))
            r25 = to_number(row.get("MAT Q2 2025_LCD MNF"))
            row_total = r23 + r24 + r25

            sector_revenue[normalize_sector(row.get("Sector"))] += row_total
            innovation_revenue[normalize_innovation(row.get("Innovation Insights"))] += row_total

            brand = row.get("International Product")
            if not brand:
                continue
            brand = str(brand).strip()
            if not brand:
                continue

            entry = brand_detail.setdefault(
                brand,
                {"r23": 0.0, "r24": 0.0, "r25": 0.0, "manufacturer": None, "corporation": None},
            )
            entry["r23"] += r23
            entry["r24"] += r24
            entry["r25"] += r25

            if not entry["manufacturer"]:
                m = row.get("Manufacturer")
                if m:
                    entry["manufacturer"] = str(m).strip()

            if not entry["corporation"]:
                c = row.get("Corporation")
                if c:
                    entry["corporation"] = str(c).strip()

        competition_count = len(brand_detail)

        brand_totals = {
            b: d["r23"] + d["r24"] + d["r25"] for b, d in brand_detail.items()
        }

        top_brand_rev = max(brand_totals.values(), default=0.0)

        total_rev_3y = (
            rev_2023 +
            rev_2024 +
            rev_2025
        )

        dominance_ratio = (
            top_brand_rev / total_rev_3y
            if total_rev_3y > 0
            else 0.0
        )

        leading_r25 = max((d["r25"] for d in brand_detail.values()), default=0.0)

        brands = []
        for brand, d in brand_detail.items():
            total = brand_totals[brand]
            brand_cagr = compute_cagr(d["r23"], d["r24"], d["r25"])
            is_new_entrant = d["r23"] == 0 and d["r25"] > 0

            trend = []
            if leading_r25 > 0 and d["r25"] == leading_r25:
                trend.append("LEADING")
            if is_new_entrant:
                trend.append("NEW_ENTRANT")
            elif brand_cagr > 20:
                trend.append("RISING")
            elif brand_cagr < -10:
                trend.append("DECLINING")
            else:
                trend.append("STABLE")

            brands.append({
                "Brand": brand,
                "Manufacturer": d["manufacturer"] or "Unknown",
                "Corporation": d["corporation"] or "Unknown",
                "Revenue_2023": round(d["r23"], 2),
                "Revenue_2024": round(d["r24"], 2),
                "Revenue_2025": round(d["r25"], 2),
                "Market_Share": round(total / total_rev_3y, 4) if total_rev_3y > 0 else 0.0,
                "Brand_CAGR": round(brand_cagr, 2),
                "Trend": trend,
            })
        brands.sort(key=lambda b: b["Market_Share"], reverse=True)

        sector_split = {
            key: round(sector_revenue.get(key, 0.0) / total_rev_3y, 4) if total_rev_3y > 0 else 0.0
            for key in SECTOR_KEYS + ["UNKNOWN"]
        }

        innovation_mix = {
            key: round(innovation_revenue.get(key, 0.0) / total_rev_3y, 4) if total_rev_3y > 0 else 0.0
            for key in INNOVATION_KEYS + ["UNKNOWN"]
        }

        MONOPOLY_THRESHOLD = 0.80

        monopoly_flag = (
            dominance_ratio >= MONOPOLY_THRESHOLD
            and competition_count > 1
        )

        # Opportunity Score: 40% revenue + 40% CAGR (revenue-weighted) + 20% competition (revenue-weighted)
        rev_norm = norm_revenue(rev_2025)
        base_score = (
            rev_norm * 0.40
            + norm_cagr(rev_cagr) * rev_norm * 0.40
            + norm_competition(competition_count) * rev_norm * 0.20
        ) * 100

        # Rank-down multipliers — apply only the most severe single penalty
        _penalties = []
        if rev_2025 == 0:
            _penalties.append(0.10)
        if rev_cagr < -20:
            _penalties.append(0.60)
        if rev_2023 > 0 and rev_2025 < rev_2023 * 0.50:
            _penalties.append(0.50)
        score = base_score * (min(_penalties) if _penalties else 1.0)

        opportunity_score = round(max(0.0, score), 1)

        # Flags
        flags = []
        if competition_count == 1:
            flags.append("SINGLE_BRAND")
        if rev_2025 == 0:
            flags.append("DEAD")
        if rev_2023 > 0 and rev_2025 == 0:
            flags.append("EXITING")
        if rev_cagr < -20:
            flags.append("COLLAPSING")
        if rev_2023 > 0 and rev_2024 > rev_2023 * 1.5 and rev_2025 < rev_2024 * 0.7:
            flags.append("SPIKE")
        if std_cagr < 0 and rev_cagr > 0:
            flags.append("VOL_DOWN_REV_UP")
        if round(dominance_ratio, 2) >= 0.60 and competition_count > 1:
            flags.append("HIGH_DOMINANCE")

        analytics.append({
            "Molecule": molecule,
            "Opportunity_Score": opportunity_score,
            "Competition_Count": competition_count,
            "Dominance_Ratio": round(dominance_ratio, 4),
            "Monopoly_Flag": monopoly_flag,
            "Revenue_2023": round(rev_2023, 2),
            "Revenue_2024": round(rev_2024, 2),
            "Revenue_2025": round(rev_2025, 2),
            "STD_2023": round(std_2023, 2),
            "STD_2024": round(std_2024, 2),
            "STD_2025": round(std_2025, 2),
            "STD_CAGR": round(std_cagr, 2),
            "Revenue_CAGR": round(rev_cagr, 2),
            "Flags": flags,
            "Brands": brands,
            "Sector_Split": sector_split,
            "Innovation_Mix": innovation_mix,
        })

    # Create two separate analyses

    # ANALYSIS 1: Before monopoly removal
    # Two-tier sort: molecules with Rev_2025 > 0 first (by STD_CAGR), then dead/zero at bottom
    analysis_1_growth = sorted(
        analytics,
        key=lambda x: (x["Revenue_2025"] == 0, -x["STD_CAGR"])
    )

    # ANALYSIS 2: After monopoly removal
    # Remove monopolies and sort by Opportunity_Score
    analysis_2_revenue = [m for m in analytics if not m["Monopoly_Flag"]]
    analysis_2_revenue.sort(key=lambda x: x["Opportunity_Score"], reverse=True)

    return {
        "analysis_1_growth": analysis_1_growth,
        "analysis_2_revenue": analysis_2_revenue,
    }


def parse_workbook(contents: bytes):
    wb = load_workbook(io.BytesIO(contents))

    ws = wb.active

    if not ws:
        raise ValueError("No sheet found in workbook")

    headers = []

    for cell in ws[1]:
        headers.append(cell.value)

    rows = []
    molecules = set()
    products = set()

    for row in ws.iter_rows(
        min_row=2,
        values_only=True
    ):
        if not any(row):
            continue

        row_dict = {}

        for col_idx, header in enumerate(headers):
            if (
                header and
                col_idx < len(row)
            ):
                row_dict[header] = row[col_idx]

        rows.append(row_dict)

        if (
            "Molecule List" in row_dict and
            row_dict["Molecule List"]
        ):
            molecules.add(
                str(row_dict["Molecule List"])
            )

        if (
            "International Product" in row_dict and
            row_dict["International Product"]
        ):
            products.add(
                str(
                    row_dict[
                        "International Product"
                    ]
                )
            )

    return rows, molecules, products


def build_upload_response(rows, molecules, products) -> Dict[str, Any]:
    analytics_result = compute_analytics(rows)

    return {
        "success": True,
        "analysis_1_growth": {
            "description": "Before monopoly removal",
            "sort_by": "STD_CAGR (Volume Growth)",
            "filter": "None (all products included)",
            "results": analytics_result["analysis_1_growth"],
            "count": len(analytics_result["analysis_1_growth"])
        },
        "analysis_2_revenue": {
            "description": "After monopoly removal",
            "sort_by": "Opportunity_Score",
            "filter": "Excluded: monopolies (80%+ dominance)",
            "results": analytics_result["analysis_2_revenue"],
            "count": len(analytics_result["analysis_2_revenue"])
        },
        "total_rows": len(rows),
        "unique_molecules": len(molecules),
        "unique_products": len(products),
    }


@app.get("/health")
async def health_check():
    return {"status": "ok"}


@app.get("/default-dataset")
async def get_default_dataset():
    if default_dataset_cache is None:
        return {
            "success": False,
            "error": "No default dataset configured on the server",
            "analytics": [],
        }

    return default_dataset_cache


@app.post("/upload")
async def upload_dataset(
    file: UploadFile = File(...)
):
    try:
        contents = await file.read()

        rows, molecules, products = parse_workbook(contents)

        return build_upload_response(rows, molecules, products)

    except Exception as e:
        return {
            "success": False,
            "error": str(e),
            "analytics": [],
        }


if __name__ == "__main__":
    import uvicorn

    port = int(
        os.getenv("PORT", "8000")
    )

    uvicorn.run(
        app,
        host="0.0.0.0",
        port=port
    )
